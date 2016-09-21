# 纯文本文件 city.txt为城市信息
#encoding=utf-8
import json
from openpyxl import Workbook
import datetime

def getJsonData():
	with open('city.txt','r') as f:
		data=json.load(f)
	return data

def writeToXls(data):
	wb=Workbook()
	ws = wb.create_sheet(title='city')
	for (k,v) in data.items():
		ws.cell(row=int(k), column=1).value = k
		ws.cell(row=int(k), column=2).value = v
	wb.save("city.xlsx")
	
if __name__ == '__main__':
	writeToXls(getJsonData())
