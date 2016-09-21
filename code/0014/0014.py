# 纯文本文件 student.txt为学生信息, 请将上述内容写到 student.xls 文件中
#encoding=utf-8
import json
from openpyxl import Workbook
import datetime

def getJsonData():
	with open('student.txt','r') as f:
		data=json.load(f)
	return data

def writeToXls(data):
	wb=Workbook()
	ws = wb.create_sheet(title='student')
	for (k,v) in data.items():
		v.insert(0,k)
		for j in range(1,len(v)+1):
					ws.cell(row=int(k), column=j).value = v[j-1]
	wb.save("sample.xlsx")
	
if __name__ == '__main__':
	writeToXls(getJsonData())
