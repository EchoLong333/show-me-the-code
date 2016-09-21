# 纯文本文件 numbers.txt, 里面的内容（包括方括号）如下所示：
#encoding=utf-8
import json
from openpyxl import Workbook
import datetime

def getJsonData():
	with open('numbers.txt','r') as f:
		data=json.load(f)
	print(data)
	return data

def writeToXls(data):
	wb=Workbook()
	ws = wb.create_sheet(title='numbers')
	for i in range(len(data)):
		for j in range(len(data[i])):
			ws.cell(row=i+1, column=j+1).value = data[i][j]
	wb.save("numbers.xlsx")
	
if __name__ == '__main__':
	writeToXls(getJsonData())
